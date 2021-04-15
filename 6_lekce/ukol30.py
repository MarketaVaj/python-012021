# Export do Excelu
# pandas umí uložit data i do Excel sešitu, se kterým se bude uživatelům lépe pracovat. K ukádání do Excelu využívá pandas modul openpyxl, který ale není nainstalován automaticky.
#
# Nainstaluj si modul openpyxl standardním způsobem, který jsme si ukazovali v lekci.
# Ulož tabulku s cekovými počty vykázaných hodin, kterou jsi vytvořila v příkladu 27 jako Excel soubor. Pokud jsi tento příklad nezpracovala, ulož libovolnou jinou tabulku jako Excel sešit. Výsledný sešit si otevři a zkontroluj. K uložení použij funkci to_excel, se kterou pracuj stejně, jako jsme na lekci pracovali s funkci to_csv.

# Použito v úkolu 27
# zamestnanciProjekty.to_excel("zamestanciProjekty.xlsx")

# Zkus data z Excelu naopak načíst pomocí funkci read_excel() a ověř, že se soubor načetl v pořádku.

# import openpyxl
# # zam = openpyxl.read_excel("zamestanciProjekty.xlsx")
# zam = openpyxl.load_workbook("zamestanciProjekty.xlsx")
# print(zam)

# import openpyxl
# from openpyxl import Workbook("zamestanciProjekty.xlsx")

from openpyxl import Workbook
zam = Workbook("zamestanciProjekty.xlsx")
print(zam)